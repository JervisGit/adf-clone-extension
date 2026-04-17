#!/usr/bin/env python3
"""
sql_runner.py  — called by sqlClient.js as a child process.

Reads one JSON command from stdin, executes it against Azure SQL
using pyodbc + AzureCliCredential, and writes one JSON result to stdout.

Command schema (all fields optional depending on operation):
{
  "server":    "test-sql-jvs.database.windows.net",
  "database":  "free-sql-db-8537067",
  "operation": "scripts" | "storedProcedure" | "bulkInsert",

  -- for "scripts":
  "scripts": [{ "type": "Query"|"NonQuery", "text": "..." }],

  -- for "storedProcedure":
  "procName": "[dbo].[usp_MyProc]",
  "params":   { "ParamName": { "type": "String", "value": "..." } },

  -- for "bulkInsert":
  "schema":   "dbo",
  "table":    "ProductsImported",
  "columns":  ["id", "product_name", ...],
  "rows":     [{ "id": "101", "product_name": "Mouse", ... }]
}

Output (written to stdout):
{
  "ok": true,
  "rowsAffected": 5,
  "recordset": [...]   -- for scripts/storedProcedure with a result set
}
or on error:
{
  "ok": false,
  "error": "message"
}
"""

import sys
import json
import struct
import re
import time
import traceback

# Lone surrogates (\uD800-\uDFFF) in cell values cause pyodbc to fail when it
# encodes strings to UTF-16LE for the ODBC driver.  Strip them before insert.
_SURROGATE_RE = re.compile(r'[\ud800-\udfff]')

def _sanitize_value(v):
    """Remove lone surrogate characters from string values."""
    return _SURROGATE_RE.sub('', v) if isinstance(v, str) else v

# Transient SQL errors that warrant a retry (Azure SQL serverless auto-pause,
# throttling, brief unavailability, etc.)
TRANSIENT_SQL_ERRORS = {40613, 40197, 40501, 49918, 49919, 49920, 4221, 233, 64}
MAX_CONNECT_RETRIES = 5
RETRY_BASE_DELAY_S  = 5   # seconds; doubles each attempt

try:
    import pyodbc
except ImportError:
    json.dump({"ok": False, "error": "pyodbc is not installed. Run: pip install pyodbc"}, sys.stdout)
    sys.exit(1)

try:
    from azure.identity import AzureCliCredential
except ImportError:
    json.dump({"ok": False, "error": "azure-identity is not installed. Run: pip install azure-identity"}, sys.stdout)
    sys.exit(1)

ODBC_DRIVERS = [
    "ODBC Driver 18 for SQL Server",
    "ODBC Driver 17 for SQL Server",
    "ODBC Driver 13 for SQL Server",
]

SQL_TOKEN_SCOPE = "https://database.windows.net/.default"
SQL_CXNATTR_TOKEN = 1256  # SQL_COPT_SS_ACCESS_TOKEN


def get_token():
    cred = AzureCliCredential()
    token = cred.get_token(SQL_TOKEN_SCOPE).token
    # JWT tokens are ASCII-only (base64url). On Windows, the az CLI subprocess
    # output is sometimes decoded with the wrong codec, introducing lone surrogates.
    # Strip any non-ASCII characters before encoding; this is always safe for JWTs.
    token = token.encode("ascii", errors="ignore").decode("ascii")
    token_bytes = token.encode("utf-16-le")
    return struct.pack(f"<I{len(token_bytes)}s", len(token_bytes), token_bytes)


def find_driver():
    available = pyodbc.drivers()
    for d in ODBC_DRIVERS:
        if d in available:
            return d
    raise RuntimeError(
        f"No supported ODBC driver found. Available: {available}. "
        f"Install 'ODBC Driver 18 for SQL Server' from https://aka.ms/odbc18"
    )


def _is_transient(exc):
    """Return True if the pyodbc error looks like a transient Azure SQL error."""
    msg = str(exc)
    for code in TRANSIENT_SQL_ERRORS:
        if f"({code})" in msg or f" {code} " in msg:
            return True
    return False


def connect(server, database):
    driver = find_driver()
    conn_str = (
        f"Driver={{{driver}}};"
        f"Server=tcp:{server},1433;"
        f"Database={database};"
        f"Encrypt=yes;TrustServerCertificate=no;"
    )
    last_exc = None
    for attempt in range(1, MAX_CONNECT_RETRIES + 1):
        try:
            token_struct = get_token()  # refresh token on each attempt
            conn = pyodbc.connect(conn_str, attrs_before={SQL_CXNATTR_TOKEN: token_struct}, timeout=60)
            conn.autocommit = False
            return conn
        except pyodbc.Error as exc:
            last_exc = exc
            if attempt < MAX_CONNECT_RETRIES and _is_transient(exc):
                delay = RETRY_BASE_DELAY_S * (2 ** (attempt - 1))  # 5, 10, 20, 40 s
                sys.stderr.write(
                    f"[sql_runner] Transient SQL error on attempt {attempt}/{MAX_CONNECT_RETRIES}, "
                    f"retrying in {delay}s: {exc}\n"
                )
                time.sleep(delay)
            else:
                raise
    raise last_exc


def rows_to_dicts(cursor):
    if cursor.description is None:
        return []
    cols = [d[0] for d in cursor.description]
    return [dict(zip(cols, [str(v) if v is not None else None for v in row]))
            for row in cursor.fetchall()]


def run_scripts(conn, scripts):
    total_affected = 0
    last_recordset = []
    for script in scripts:
        text = (script.get("text") or "").strip()
        if not text:
            continue
        cursor = conn.cursor()
        cursor.execute(text)
        try:
            last_recordset = rows_to_dicts(cursor)
        except pyodbc.ProgrammingError:
            last_recordset = []
        affected = cursor.rowcount if cursor.rowcount >= 0 else 0
        total_affected += affected
    conn.commit()
    return {"ok": True, "rowsAffected": total_affected, "recordset": last_recordset}


def run_stored_procedure(conn, proc_name, params):
    cursor = conn.cursor()
    # Build EXEC call with positional parameters
    param_placeholders = ", ".join(["?"] * len(params))
    sql = f"EXEC {proc_name} {param_placeholders}" if params else f"EXEC {proc_name}"
    values = [p.get("value") for p in params.values()]
    cursor.execute(sql, values)
    try:
        recordset = rows_to_dicts(cursor)
    except pyodbc.ProgrammingError:
        recordset = []
    affected = cursor.rowcount if cursor.rowcount >= 0 else 0
    conn.commit()
    return {"ok": True, "rowsAffected": affected, "recordset": recordset}


def run_bulk_insert(conn, schema, table, columns, rows):
    if not rows:
        return {"ok": True, "rowsAffected": 0, "recordset": []}
    qualified = f"[{schema}].[{table}]"
    col_list  = ", ".join(f"[{c}]" for c in columns)
    placeholders = ", ".join(["?"] * len(columns))
    sql = f"INSERT INTO {qualified} ({col_list}) VALUES ({placeholders})"
    cursor = conn.cursor()
    data = [[_sanitize_value(row.get(c)) for c in columns] for row in rows]
    cursor.fast_executemany = True
    cursor.executemany(sql, data)
    affected = len(data)
    conn.commit()
    return {"ok": True, "rowsAffected": affected, "recordset": []}


def read_parquet(file_path):
    """Read a local parquet file and return rows as a list of dicts.
    Does not require a database connection."""
    try:
        import pyarrow.parquet as pq
    except ImportError:
        return {"ok": False, "error": "pyarrow not installed. Run: pip install pyarrow"}
    table = pq.read_table(file_path)
    cols = table.column_names
    col_data = {c: table.column(c).to_pylist() for c in cols}
    rows = []
    for i in range(len(table)):
        row = {}
        for c in cols:
            v = col_data[c][i]
            row[c] = str(v) if v is not None else None
        rows.append(row)
    return {"ok": True, "rows": rows, "columns": cols}


def read_excel(file_path, sheet_name=None, first_row_as_header=True, null_value=""):
    """Read an Excel (.xlsx) file using openpyxl. Does not require a DB connection."""
    try:
        import openpyxl
    except ImportError:
        return {"ok": False, "error": "openpyxl not installed. Run: pip install openpyxl"}
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    # Collect non-empty rows
    data = [row for row in ws.iter_rows(values_only=True)
            if any(v is not None for v in row)]
    if not data:
        return {"ok": True, "rows": [], "columns": []}

    def to_str(v):
        if v is None:
            return null_value
        return str(v)

    if first_row_as_header:
        columns = [str(c) if c is not None else f"col{i}" for i, c in enumerate(data[0])]
        rows = [dict(zip(columns, [to_str(v) for v in row])) for row in data[1:]]
    else:
        columns = [f"col{i}" for i in range(len(data[0]))]
        rows = [dict(zip(columns, [to_str(v) for v in row])) for row in data]
    return {"ok": True, "rows": rows, "columns": columns}


def read_xml(file_path, row_tag=None, null_value=""):
    """Read an XML file using lxml. Each child of root (or elements matching row_tag)
    becomes a row; element children become columns. Does not require a DB connection."""
    try:
        from lxml import etree
    except ImportError:
        return {"ok": False, "error": "lxml not installed. Run: pip install lxml"}
    try:
        tree = etree.parse(file_path)
    except etree.XMLSyntaxError as e:
        return {"ok": False, "error": f"XML parse error: {e}"}
    root = tree.getroot()
    if row_tag:
        row_elements = root.findall(f".//{row_tag}")
        if not row_elements:
            local = etree.QName(row_tag).localname
            row_elements = [el for el in root.iter()
                            if etree.QName(el.tag).localname == local]
    else:
        row_elements = list(root)
    if not row_elements:
        return {"ok": True, "rows": [], "columns": []}
    # Union of column names (preserve first-seen order)
    col_order = {}
    for el in row_elements:
        for attr in el.attrib:
            col_order[f"@{attr}"] = True
        for child in el:
            col_order[etree.QName(child.tag).localname] = True
    columns = list(col_order.keys())
    rows = []
    for el in row_elements:
        row = {c: null_value for c in columns}
        for attr, val in el.attrib.items():
            row[f"@{attr}"] = val if val is not None else null_value
        for child in el:
            tag = etree.QName(child.tag).localname
            row[tag] = (child.text.strip() if child.text else null_value)
        rows.append(row)
    return {"ok": True, "rows": rows, "columns": columns}


def main():
    raw = sys.stdin.read()
    cmd = json.loads(raw)
    operation = cmd.get("operation", "scripts")

    # File-read operations — do not require a database connection
    if operation in ("readParquet", "readExcel", "readXml"):
        if operation == "readParquet":
            result = read_parquet(cmd["filePath"])
        elif operation == "readExcel":
            result = read_excel(
                cmd["filePath"],
                sheet_name=cmd.get("sheetName"),
                first_row_as_header=cmd.get("firstRowAsHeader", True),
                null_value=cmd.get("nullValue", ""),
            )
        elif operation == "readXml":
            result = read_xml(
                cmd["filePath"],
                row_tag=cmd.get("rowTag"),
                null_value=cmd.get("nullValue", ""),
            )
        json.dump(result, sys.stdout)
        return

    server    = cmd["server"]
    database  = cmd["database"]

    conn = connect(server, database)
    try:
        if operation == "scripts":
            result = run_scripts(conn, cmd.get("scripts", []))
        elif operation == "storedProcedure":
            result = run_stored_procedure(conn, cmd["procName"], cmd.get("params", {}))
        elif operation == "bulkInsert":
            result = run_bulk_insert(
                conn,
                cmd.get("schema", "dbo"),
                cmd["table"],
                cmd.get("columns", []),
                cmd.get("rows", []),
            )
        else:
            result = {"ok": False, "error": f"Unknown operation: {operation}"}
    finally:
        conn.close()

    json.dump(result, sys.stdout)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        json.dump({"ok": False, "error": str(e), "trace": traceback.format_exc()}, sys.stdout)
        sys.exit(1)
